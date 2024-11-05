# verificar se valores estão pagos ou atrasados , e caso estejam pagos , informar metodo de pagamento
# Com base na planilha pegar cpf  e consultar no site disponibilizado para verificar  se a conta está(ou não) paga
# Caso esteja paga, preencher a planilha de fechamento com um "OK", caso contrario, informar que está "pendente"

# -Passos Manuais, para transformar em código
# Passo 1- Abrir a Planilha , copiar cpf do cliente
# Passo 2- Abrir o site, colar o cpf do cliente e clicar em consultar
# Passo 3- Verificar se está em dia ou atrasado
# Passo 4- se estiver em dia, pegar data do pagamento e forma de pagamento e  prencher planilha com "OK".
# Passo 5- se estiver atrasado, colocar na planilha status como pendente
# Passo 6 - Inserir essas novas informações (nome,valor,cpf,vencimento,status,e caso esteja em dia, data de pagamento e metodo de pagamento(cartão, boleto)
# em uma nova planilha)
# Passo 7- Repetir até o último cliente


import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
# Passo 1- Abrir a Planilha , copiar cpf do cliente

planilha_clientes = openpyxl.load_workbook("dados_clientes.xlsx")
pagina_cliente = planilha_clientes['Sheet1']

for linha in pagina_cliente.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha
# Passo 2- Abrir o site, colar o cpf do cliente e clicar em consultar
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get('https://consultcpf-devaprender.netlify.app/')
    sleep(5)
    campo_pesquisa_cpf = driver.find_element(
        By.XPATH, "//input[@id='cpfInput']")
    sleep(1)
    campo_pesquisa_cpf.clear()
    campo_pesquisa_cpf.send_keys(cpf)
    sleep(1)
    campo_consultar = driver.find_element(By.XPATH, "//button[@type='submit']")
    sleep(1)
    campo_consultar.click()
    sleep(4)
    # Passo 3- Verificar se está em dia ou atrasado
    status=driver.find_element(By.XPATH, "//span[@id='statusLabel']")
    # Passo 4- se estiver em dia, pegar data do pagamento e forma de pagamento e  prencher planilha com "OK".
    if status.text == 'em dia':
        data_pagamento = driver.find_element(
            By.XPATH, "//p[@id='paymentDate']")
        metodo_pagamento = driver.find_element(
            By.XPATH, "//p[@id='paymentMethod']")
        data_pagamento_limpo = data_pagamento.text.split()[3]
        metodo_pagamento_limpo = metodo_pagamento.text.split()[3]
        planilha_fechamento = openpyxl.load_workbook(
            'planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']
        pagina_fechamento.append(
            [nome, valor, cpf, vencimento, 'em dia', data_pagamento_limpo, metodo_pagamento_limpo])
        planilha_fechamento.save('planilha fechamento.xlsx')
    else:
        planilha_fechamento = openpyxl.load_workbook(
            'planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']
        pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])
        planilha_fechamento.save('planilha fechamento.xlsx')
    # Passo 5- se estiver atrasado, colocar na planilha status como pendente
